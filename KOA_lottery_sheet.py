#Lottery sheet
#Daniel Eberhart
#5/31/2020

from tkinter import *
import webbrowser

#For Excel
from openpyxl import Workbook, load_workbook

#For obtaining current time
import time

class Lottery: #Use master for Tk functions and commands
    def __init__(self, master):
        self.master = master
        master.title('Lottery Sheet')

        ##### Create frames
        self.f1 = Frame(master)
        self.f1.pack(side=LEFT)
        self.f2 = Frame(master)
        self.f2.pack(side=RIGHT)

        ##### Get current date and create a string that symbolizes the =DATE formula in Excel
        self.local_time = time.localtime(time.time())
        self.curr_time =  str(self.local_time[1]) + "/" + str(self.local_time[2]) + "/" + str(self.local_time[0]) 

        ##### Making the file menu at the top of the application
        menu = Menu(master)
        master.config(menu=menu)
        menu.add_cascade(label='About', command=self.aboutWindow)
        menu.add_cascade(label='Help/FAQ', command=self.helpWindow)

        ##### Initialize inventory
        self.inventory={}
        self.checkInventory()

        ##### Make 25 Labels
        self.createLabels()

        ##### Make 25 Entries and store in dictionary 
        self.d={}            
        self.createEntries()
        
        infoLabel = Label(self.f1, text="Only enter numbers into the entry boxes please!")
        infoLabel.grid(row=26,column=0,columnspan=2)
        
        ##### Display image in top right
        photo = PhotoImage(file="oregon-lottery-logo.png")
        w = Label(self.f2, image=photo)
        w.photo = photo
        w.grid(row=0,column=0,sticky=N+E)

        ##### Buttons for when a lottery is out of stock or out
        out_btn = Button(self.f2, height=2, bg="green", text="Click this button if a lottery is out", font="bold", command=self.lotteryOut)
        out_btn.grid(row=1,column=0,columnspan=2,sticky=W+E+N+S)
        stock_btn = Button(self.f2, height=2, bg="Green", text="Click this button to restock a lottery", font="bold", command=self.restockPopup)
        stock_btn.grid(row=2,column=0,columnspan=2,sticky=W+E+N+S)

        ##### Assigning function to buttons  
        submitButton = Button(self.f2, text="Submit", bg="yellow", height=2, font="bold", command=lambda:[self.export_to_excel(), self.update_amt_sold(), master.quit()])
        submitButton.grid(row=3,column=0,columnspan=2,sticky=W+E+N+S) 
        quitButton = Button(self.f2, text="Quit", bg="red", height=2, font="bold", command=master.quit)
        quitButton.grid(row=4,column=0,columnspan=2,sticky=W+E+N+S)
      
    def createLabels(self):
        for i in range(1,25):
           lbltxt = "Lottery #" + str(i) + ":"
           l1 = Label(self.f1, text=lbltxt, font="bold")
           l1.grid(row=i, sticky=W)

    def createEntries(self):
        wb = load_workbook("Lottery-Excel.xlsx")
        ws = wb["Inventory"]
        for x in range(1,25):
            self.d[f'e{x}'] = Entry(self.f1, bg="yellow")
            self.d[f'e{x}'].grid(row=x, column=1)
            if ws.cell(row=x, column=3).value == "out":
                self.d[f'e{x}'].config(state='disabled')
        wb.save("Lottery-Excel.xlsx")

    def popUpConstructor(self, popup, w, h):
        ##### Display the new window in the middle of the screen with the appropriate dimensions
        sw = popup.winfo_screenwidth()
        sh = popup.winfo_screenheight()
        x = (sw - w)/2
        y = (sh - h)/2
        popup.geometry('%dx%d+%d+%d' % (w,h,x,y))
    
    def aboutWindow(self):
        popup = Tk()
        self.popUpConstructor(popup, 400, 200)
        popup.title("About")

        ##### Display appropriate info and contact information
        font = ("Helvetica", "15", "bold")
        msg = "Created By: Daniel Eberhart\nEmail: danieleberhart14@gmail.com"
        label = Label(popup, text=msg, width = 120, height=10, bg="yellow", font=font)
        label.pack()

    def helpWindow(self):
        popup = Tk()
        self.popUpConstructor(popup, 700, 200)
        popup.title("Help/FAQ")

        font = ("Helvetica", "12", "bold")
        msg = """Q: I already submitted the lottery numbers, but then someone came in
         and bought more. What do I do?\nA: Just re-enter all the numbers again. Unfortunately
         it is not possible to enter a number for a single lottery and change it."""
        label = Label(popup, text=msg, width = 120, height=10, bg="yellow", font=font)
        label.pack()

    def export_to_excel(self):
        ##### Initialize variables
        row_num = 1
        col_num = 2
        date_found = False

        ##### Load workbook and open selected sheet from said workbook
        wb = load_workbook("Lottery-Excel.xlsx")
        ws = wb["Data"]
        ws2 = wb["Inventory"]

        ##### Get date
        while date_found == False:
            if ws.cell(row=row_num, column=1).value == str(self.curr_time):
                #print("Date Successfully found at row " + str(row_num))
                date_found = True
            else:
                row_num += 1
    
        ##### Writing to the cells
        for x in range(1,25):
            temp_num = self.d[f'e{x}'].get()
            entry_state = self.d[f'e{x}'].cget('state')
            if entry_state == "normal":
                ws.cell(row=row_num, column=col_num, value=int(temp_num))
            else:
                inv_num = ws2.cell(row=x, column=2).value
                ws.cell(row=row_num, column=col_num, value=inv_num)
            
            col_num += 1
        ##### Save Workbook    
        wb.save("Lottery-Excel.xlsx")

    def update_amt_sold(self):
        ##### Initialize variables
        col_num = 2
        row_num = 1
        date_found = False

        ##### Load workbook and open selected sheets from said workbook
        wb = load_workbook("Lottery-Excel.xlsx")
        ws1 = wb["Sold"]
        ws2 = wb["Data"]
        

        ##### Get date
        while date_found == False:
            if ws1.cell(row=row_num, column=1).value == str(self.curr_time):
                date_found = True
            else:
                row_num += 1

        ##### Assign sold data and calculate total sold
        total_sold = 0
        for i in range(1,25):
            minuend = ws2.cell(row=row_num, column=col_num).value
            subtrahend = ws2.cell(row=row_num-1, column=col_num).value
            difference = int(minuend) - int(subtrahend)

            if difference < 0:
                difference = ws2.cell(row=row_num, column=col_num).value

            total_sold += difference
            ws1.cell(row=row_num, column=col_num, value=int(difference))
            col_num += 1

        ##### Assign the total sold for today's date    
        ws1.cell(row=row_num, column=col_num, value=int(total_sold))

        ##### Save Workbook
        wb.save("Lottery-Excel.xlsx")

    def checkInventory(self):
        row_num = 1
        wb = load_workbook("Lottery-Excel.xlsx")
        ws = wb["Inventory"] 
        ##### Load inventory
        for x in range(1,25):
            self.inventory[f'inv{x}'] = ws.cell(row=row_num, column=3).value
            row_num += 1
        wb.save("Lottery-Excel.xlsx")

    def restockPopup(self):
        popup = Tk()
        self.popUpConstructor(popup, 260, 75)
        popup.title("Restock")

        label2 = Label(popup, text="Lottery # to restock (1-24): ")
        label2.grid(row=1,column=0, sticky=W)
        self.restock_num = Entry(popup, width=5)
        self.restock_num.grid(row=1, column=1)

        label3 = Label(popup, text="How many does the new lottery set have: ")
        label3.grid(row=2,column=0, sticky=W)
        self.restock_amt = Entry(popup, width=5)
        self.restock_amt.grid(row=2, column=1)

        btn_1 = Button(popup, text="Submit", command=lambda:[self.restockInventory(), popup.destroy()])
        btn_1.grid(row=3, rowspan=2, column=0, columnspan=3)

    def restockInventory(self):
        wb = load_workbook("Lottery-Excel.xlsx")
        ws = wb["Inventory"]
        lottery_num = self.restock_num.get()
        new_inventory_value = self.restock_amt.get()
        if int(lottery_num) <= 24 and int(lottery_num) >= 1:
            ws.cell(row=int(lottery_num), column=2, value=int(new_inventory_value))
            ws.cell(row=int(lottery_num), column=3, value="in")
            self.d[f'e{int(lottery_num)}'].config(state='normal')

        else:
            self.restockPopup()
        wb.save("Lottery-Excel.xlsx")

    def lotteryOut(self):
        popup = Tk()
        self.popUpConstructor(popup, 260, 75)
        popup.title("Restock")
        msg = "Please indicate which lottery # is out."
        label1 = Label(popup, text=msg)
        label1.grid(row=0,column=0,columnspan=3) 

        label2 = Label(popup, text="Lottery # that is out (1-24): ")
        label2.grid(row=1,column=0, sticky=W)
        self.inout_bool = Entry(popup, width=5)
        self.inout_bool.grid(row=1, column=1)

        btn_1 = Button(popup, text="Submit", command=lambda:[self.assignOut(), popup.destroy()])
        btn_1.grid(row=3, rowspan=2, column=0, columnspan=3)

    def assignOut(self):
        wb = load_workbook("Lottery-Excel.xlsx")
        ws = wb["Inventory"]
        lottout = self.inout_bool.get()
        if int(lottout) <= 24 and int(lottout) >= 1:
            ws.cell(row=int(lottout), column=3, value="out")
            self.d[f'e{int(lottout)}'].config(state='disabled')
        else:
            self.lotteryOut()
        wb.save("Lottery-Excel.xlsx")


root = Tk()
mygui = Lottery(root)
root.mainloop()
